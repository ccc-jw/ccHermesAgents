from pathlib import Path

from openpyxl import load_workbook

from scripts.generate_test_checklist import generate_checklist


def test_generate_checklist_creates_excel(tmp_path: Path):
    source = tmp_path / "cases.md"
    output = tmp_path / "checklist.xlsx"
    source.write_text("## 登录成功\n步骤: 输入账号密码\n预期: 返回 token\n", encoding="utf-8")

    generate_checklist(source, output)

    workbook = load_workbook(output)
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "序号"
    assert sheet.cell(row=2, column=2).value == "登录成功"
